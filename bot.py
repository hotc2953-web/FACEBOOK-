import os, re, json
from openpyxl import Workbook, load_workbook
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters

# ========== CONFIG ==========
TOKEN = "8190004924:AAEPwuTN_yTpnjvzoOsxdrTQ5VrXSjTX9AM"
USERS_FILE = "users.json"
DATA_DIR = "user_files"

os.makedirs(DATA_DIR, exist_ok=True)

# ========== USERS ==========
def load_users():
    if not os.path.exists(USERS_FILE):
        return {}
    with open(USERS_FILE, "r") as f:
        return json.load(f)

def save_users(d):
    with open(USERS_FILE, "w") as f:
        json.dump(d, f, indent=2)

users = load_users()

# ========== UTILS ==========
def extract_uid(cookie):
    m = re.search(r'c_user\s*=\s*(\d{6,25})', cookie)
    return m.group(1) if m else ""

def user_file(tg_uid):
    return os.path.join(DATA_DIR, f"current_{tg_uid}.xlsx")

def create_new_file(path):
    wb = Workbook()
    ws = wb.active
    ws.append(["UID", "PASSWORD", "COOKIE"])
    wb.save(path)

def get_sheet(path):
    if not os.path.exists(path):
        create_new_file(path)
    wb = load_workbook(path)
    return wb, wb.active

# ========== COMMANDS ==========
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = str(update.effective_user.id)
    if uid not in users:
        users[uid] = {"pass": None}
        save_users(users)

    msg = (
        "🔐 Password set karo:\n"
        "/set MANISH@19\n\n"
        "Password set hone ke baad sirf cookie paste karo.\n"
        "Saari cookie ek hi file me save hongi.\n\n"
        "📂 Latest file:\n/latest\n\n"
        "🆕 Nayi file start:\n/reset"
    )
    await update.message.reply_text(msg)

async def setpass(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = str(update.effective_user.id)
    if not context.args:
        await update.message.reply_text("Usage: /set MANISH@18")
        return

    users.setdefault(uid, {})
    users[uid]["pass"] = context.args[0]
    save_users(users)

    await update.message.reply_text(
        "✅ Password set ho gaya\nAb cookie paste karo"
    )

async def changepass(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = str(update.effective_user.id)
    if uid not in users or not users[uid].get("pass"):
        await update.message.reply_text("❌ Pehle password set karo")
        return

    if not context.args:
        await update.message.reply_text("Usage: /change NEWPASS@123")
        return

    users[uid]["pass"] = context.args[0]
    save_users(users)
    await update.message.reply_text("🔁 Password change ho gaya")

async def reset_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = str(update.effective_user.id)
    path = user_file(uid)
    create_new_file(path)
    await update.message.reply_text("🆕 New file start ho gayi")

async def latest_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = str(update.effective_user.id)
    path = user_file(uid)
    if not os.path.exists(path):
        await update.message.reply_text("❌ Abhi koi file nahi hai")
        return

    await update.message.reply_document(
        document=open(path, "rb"),
        caption="📂 Latest file"
    )

# ========== COOKIE HANDLER ==========
async def cookie_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = str(update.effective_user.id)
    text = update.message.text.strip()

    if uid not in users or not users[uid].get("pass"):
        await update.message.reply_text("❌ Pehle password set karo")
        return

    path = user_file(uid)
    wb, ws = get_sheet(path)

    added = 0
    for cookie in text.splitlines():
        fb_uid = extract_uid(cookie)
        if not fb_uid:
            continue
        ws.append([fb_uid, users[uid]["pass"], cookie])
        added += 1

    if added == 0:
        await update.message.reply_text("⚠️ Cookie me UID nahi mila")
        return

    wb.save(path)

    await update.message.reply_text(f"✅ {added} cookie save hui")
    await update.message.reply_document(
        document=open(path, "rb"),
        caption="📂 Updated latest file"
    )

# ========== APP ==========
app = ApplicationBuilder().token(TOKEN).build()
app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("set", setpass))
app.add_handler(CommandHandler("change", changepass))
app.add_handler(CommandHandler("reset", reset_file))
app.add_handler(CommandHandler("latest", latest_file))
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, cookie_handler))

print("Bot running...")
app.run_polling()